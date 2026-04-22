from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = ROOT / "data" / "raw"
NORMALIZED_DIR = ROOT / "data" / "normalized"

VOCAB_SOURCE = RAW_DIR / "vocabbbb.xlsx"
KANJI_SOURCE = RAW_DIR / "tabla-de-kanji-genki.xlsx"

VOCAB_OUTPUT = NORMALIZED_DIR / "genki-vocab.json"
KANJI_OUTPUT = NORMALIZED_DIR / "genki-kanji.json"


def compact(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def parse_lesson(value: Any) -> int:
    text = compact(value)
    if text.lower().startswith("l"):
        text = text[1:]
    return int(text)


def make_id(prefix: str, lesson: int, *parts: str) -> str:
    basis = "|".join([str(lesson), *parts]).encode("utf-8")
    digest = hashlib.sha1(basis).hexdigest()[:12]
    return f"{prefix}-l{lesson:02d}-{digest}"


def split_alt_meanings(meaning_es: str) -> list[str]:
    parts = [part.strip() for part in meaning_es.split(";")]
    unique_parts: list[str] = []
    for part in parts:
        if part and part not in unique_parts:
            unique_parts.append(part)
    return unique_parts if len(unique_parts) > 1 else []


def contains_hiragana(text: str) -> bool:
    return any("\u3040" <= char <= "\u309f" for char in text)


BARE_NUMERAL_PATTERN = re.compile(r"^[一二三四五六七八九十百千]$")
CLOCK_HOUR_PATTERN = re.compile(r"^[一二三四五六七八九十]時$")


def classify_kanji_entry(written_form: str, kana_reading: str, meaning_es: str) -> str:
    if "さん" in written_form:
        return "proper_name"
    if written_form.endswith("曜日"):
        return "weekday"
    if CLOCK_HOUR_PATTERN.match(written_form):
        return "time_expression"
    if written_form.endswith("する"):
        return "suru_verb"
    if written_form.endswith("な"):
        return "na_adjective"
    if "の" in written_form:
        return "phrase"
    if len(written_form) == 1:
        return "single_kanji_word"
    if contains_hiragana(written_form) and written_form.endswith("い"):
        return "adjective_i"
    if contains_hiragana(written_form) and kana_reading.endswith(("う", "く", "ぐ", "す", "つ", "ぬ", "ぶ", "む", "る")):
        return "verb"
    return "word"


def should_keep_kanji_entry(written_form: str, entry_type: str) -> bool:
    if entry_type == "proper_name":
        return False
    return True


def build_vocab_payload() -> dict[str, Any]:
    workbook = load_workbook(VOCAB_SOURCE, data_only=True)
    lessons: dict[str, list[dict[str, Any]]] = {}
    items: list[dict[str, Any]] = []

    for sheet in workbook.worksheets:
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            lesson_value, hiragana, kanji, meaning = row
            lesson = parse_lesson(lesson_value)
            kana_reading = compact(hiragana)
            kanji_form = compact(kanji)
            meaning_es = compact(meaning)
            japanese_display = kanji_form or kana_reading
            item = {
                "id": make_id("vocab", lesson, japanese_display, kana_reading, meaning_es),
                "lesson": lesson,
                "japanese_display": japanese_display,
                "kana_reading": kana_reading,
                "kanji_form": kanji_form or None,
                "meaning_es": meaning_es,
                "alt_meanings": split_alt_meanings(meaning_es),
                "source_sheet": sheet.title,
                "source_row": row_index,
            }
            items.append(item)
            lessons.setdefault(str(lesson), []).append(item)

    return {
        "schema_version": 1,
        "source_file": str(VOCAB_SOURCE.relative_to(ROOT)),
        "module": "vocabulario",
        "lessons": lessons,
        "items": items,
    }


def build_kanji_payload() -> dict[str, Any]:
    workbook = load_workbook(KANJI_SOURCE, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    lessons: dict[str, list[dict[str, Any]]] = {}
    items: list[dict[str, Any]] = []

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        lesson_value, written, reading, meaning = row
        lesson = parse_lesson(lesson_value)
        written_form = compact(written)
        kana_reading = compact(reading)
        meaning_es = compact(meaning)
        entry_type = classify_kanji_entry(written_form, kana_reading, meaning_es)
        if not should_keep_kanji_entry(written_form, entry_type):
            continue
        item = {
            "id": make_id("kanji", lesson, written_form, kana_reading, meaning_es),
            "lesson": lesson,
            "written_form": written_form,
            "kana_reading": kana_reading,
            "meaning_es": meaning_es,
            "entry_type": entry_type,
            "is_single_kanji": len(written_form) == 1,
            "contains_okurigana": contains_hiragana(written_form),
            "source_sheet": sheet.title,
            "source_row": row_index,
        }
        items.append(item)
        lessons.setdefault(str(lesson), []).append(item)

    return {
        "schema_version": 1,
        "source_file": str(KANJI_SOURCE.relative_to(ROOT)),
        "module": "kanji",
        "lessons": lessons,
        "items": items,
    }


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    write_json(VOCAB_OUTPUT, build_vocab_payload())
    write_json(KANJI_OUTPUT, build_kanji_payload())


if __name__ == "__main__":
    main()
